import sys
import mysql.connector
import pandas as pd
import os
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(mes, anio):
    try:
        conexion = mysql.connector.connect(host="localhost", user="root", password="", database="ecobric_db")
        cursor = conexion.cursor(dictionary=True)
        
        query_ingresos = """
            SELECT p.id as Pedido_ID, u.nombre as Cliente, p.monto_total as Ingreso_EUR, p.creado_en as Fecha
            FROM pedidos p JOIN usuarios u ON p.usuario_id = u.id
            WHERE p.estado = 'PAGADO' AND MONTH(p.creado_en) = %s AND YEAR(p.creado_en) = %s
        """
        cursor.execute(query_ingresos, (mes, anio))
        ingresos = cursor.fetchall()
        
        query_gastos = """
            SELECT mi.id as Movimiento_ID, pr.nombre as Producto, mi.cantidad as Unidades, 
                   (SELECT precio_suministro FROM producto_proveedor pp WHERE pp.producto_id = mi.producto_id LIMIT 1) as Costo_Unitario_EUR,
                   (SELECT prov.nombre_empresa FROM proveedores prov JOIN producto_proveedor pp ON prov.id = pp.proveedor_id WHERE pp.producto_id = mi.producto_id LIMIT 1) as Nombre_Proveedor,
                   mi.fecha_movimiento as Fecha_Compra
            FROM movimientos_inventario mi JOIN productos pr ON mi.producto_id = pr.id
            WHERE mi.tipo_movimiento = 'ENTRADA' AND mi.notas != 'Inventario inicial'
            AND MONTH(mi.fecha_movimiento) = %s AND YEAR(mi.fecha_movimiento) = %s
        """
        cursor.execute(query_gastos, (mes, anio))
        gastos_raw = cursor.fetchall()
        
        gastos = []
        for g in gastos_raw:
            precio = float(g['Costo_Unitario_EUR']) if g['Costo_Unitario_EUR'] else 0.0
            proveedor = g['Nombre_Proveedor'] if g['Nombre_Proveedor'] else 'No Asignado'
            gastos.append({
                'Ref_Movimiento': f"MOV-{g['Movimiento_ID']}",
                'Material_Reabastecido': g['Producto'],
                'Proveedor': proveedor,
                'Unidades': g['Unidades'],
                'Costo_Unitario_EUR': round(precio, 2),
                'Total_Gasto_EUR': round(g['Unidades'] * precio, 2),
                'Fecha_Compra': g['Fecha_Compra']
            })
            
        df_ing = pd.DataFrame(ingresos) if ingresos else pd.DataFrame(columns=['Pedido_ID', 'Cliente', 'Ingreso_EUR', 'Fecha'])
        df_gas = pd.DataFrame(gastos) if gastos else pd.DataFrame(columns=['Ref_Movimiento', 'Material_Reabastecido', 'Proveedor', 'Unidades', 'Costo_Unitario_EUR', 'Total_Gasto_EUR', 'Fecha_Compra'])
        
        tot_ing = float(df_ing['Ingreso_EUR'].sum()) if not df_ing.empty else 0.0
        tot_gas = float(df_gas['Total_Gasto_EUR'].sum()) if not df_gas.empty else 0.0
        
        df_res = pd.DataFrame([{
            'Periodo Fiscal': f"{mes}/{anio}",
            'Total Ingresos (EUR)': tot_ing,
            'Total Gastos (EUR)': tot_gas,
            'Balance Neto (EUR)': tot_ing - tot_gas
        }])
        
        script_dir = os.path.dirname(os.path.abspath(__file__))
        filename = f"Reporte_Financiero_{mes}_{anio}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(script_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            sheet_name = "Reporte Maestro"
            
            # 1. ESCRIBIR RESUMEN CONTABLE
            df_res.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False)
            
            # 2. ESCRIBIR INGRESOS (Dejando 3 filas vacías)
            start_ingresos = len(df_res) + 3
            df_ing.to_excel(writer, sheet_name=sheet_name, startrow=start_ingresos, index=False)
            
            # 3. ESCRIBIR GASTOS (Dejando 3 filas vacías tras ingresos)
            start_gastos = start_ingresos + len(df_ing) + 4
            df_gas.to_excel(writer, sheet_name=sheet_name, startrow=start_gastos, index=False)
            
            # RECUPERAR WORKSHEET PARA FORMATEO
            worksheet = writer.sheets[sheet_name]
            border_style = Side(border_style="thin", color="DDDDDD")
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            
            # Títulos manuales que expliquen cada sección
            ws_title_ing = worksheet.cell(row=start_ingresos, column=1, value=">>> DETALLE DE INGRESOS (VENTAS)")
            ws_title_ing.font = Font(bold=True, color="2E7D32")
            
            ws_title_gas = worksheet.cell(row=start_gastos, column=1, value=">>> DETALLE DE GASTOS (COMPRAS A PROVEEDOR)")
            ws_title_gas.font = Font(bold=True, color="D32F2F")
            
            # Formatear el Resumen
            for col_idx in range(len(df_res.columns)):
                cell_h = worksheet.cell(row=1, column=col_idx+1)
                cell_h.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                cell_h.font = Font(color='FFFFFF', bold=True)
                cell_h.alignment = Alignment(horizontal='center')
                cell_h.border = border
                
                cell_d = worksheet.cell(row=2, column=col_idx+1)
                cell_d.alignment = Alignment(horizontal='center')
                cell_d.border = border
                if 'EUR' in df_res.columns[col_idx]:
                    cell_d.number_format = '#,##0.00 €'
                    if 'Balance' in df_res.columns[col_idx]:
                        val = float(df_res.iloc[0, col_idx])
                        cell_d.font = Font(bold=True, color='2E7D32' if val >= 0 else 'D32F2F')

            # Formatear Ingresos
            for col_idx in range(len(df_ing.columns)):
                cell_h = worksheet.cell(row=start_ingresos+1, column=col_idx+1)
                cell_h.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
                cell_h.font = Font(color='FFFFFF', bold=True)
                cell_h.border = border
                
                for r_idx in range(len(df_ing)):
                    cell_d = worksheet.cell(row=start_ingresos+2+r_idx, column=col_idx+1)
                    cell_d.border = border
                    if 'EUR' in df_ing.columns[col_idx]:
                        cell_d.number_format = '#,##0.00 €'

            # Formatear Gastos
            for col_idx in range(len(df_gas.columns)):
                cell_h = worksheet.cell(row=start_gastos+1, column=col_idx+1)
                cell_h.fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
                cell_h.font = Font(color='FFFFFF', bold=True)
                cell_h.border = border
                
                for r_idx in range(len(df_gas)):
                    cell_d = worksheet.cell(row=start_gastos+2+r_idx, column=col_idx+1)
                    cell_d.border = border
                    if 'EUR' in df_gas.columns[col_idx]:
                        cell_d.number_format = '#,##0.00 €'

            # Auto-ancho columnas asumiendo maximos aproximados
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 20
            worksheet.column_dimensions['G'].width = 20

        print(filepath)
    except Exception as e:
        print(f"Error: {str(e)}")
    finally:
        if 'conexion' in locals() and conexion.is_connected():
            cursor.close()
            conexion.close()

if __name__ == "__main__":
    mes = date.today().month if len(sys.argv) < 2 else sys.argv[1]
    anio = date.today().year if len(sys.argv) < 3 else sys.argv[2]
    export_to_excel(str(mes).zfill(2), str(anio))
